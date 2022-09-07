import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from datetime import date, timedelta
import time
import math


#  Функция для анализа входного файла и формирования нового файла (отчета)
def analyze_and_insert(excel, result_name):
    #  Исходные данные для дальнейшей работы (константы)
    articles_1 = ("650", "550", "500")
    articles_2 = ("330/1", "330")
    articles_3 = ("150", "200", "300")
    num = (13, 24, 36)

    #  Формируем словрь с датами.
    dates = dict()
    for i in range(len(excel)):
        if type(excel.iat[i, 0]) == str and (excel.iat[i, 0][-7:-1] + "0") == "0:00:00":
            dates[excel.iat[i, 0]] = i

    #  Поиск индекса столбца с заголовком Номенклатура, Итого и Склад Р
    index_of_product_j = 0
    index_of_end = 0
    index_of_werehous_j = 0
    index_of_werehous_i = 0
    for i in range(excel.shape[0]):
        for j in range(excel.shape[1]):
            if excel.iat[i, j] == "Номенклатура":
                index_of_product_j = j
                break
    for i in range(excel.shape[0]):
        if excel.iat[i, 0] == "Итого":
            index_of_end = i
            break
    for i in range(excel.shape[0]):
        for j in range(excel.shape[1]):
            if type(excel.iat[i, j]) == str and excel.iat[i, j][:7] == "Склад Р":
                index_of_werehous_i = i
                index_of_werehous_j = j
                break
    # print(str(excel.iat[index_of_werehous_i, index_of_werehous_j]))
    msg1 = str(excel.iat[index_of_werehous_i, index_of_werehous_j])
    msg2 ="C " + str(list(dates.keys())[0])[:10] + " по " + str(list(dates.keys())[-1])[:10]
    # print("C " + str(list(dates.keys())[0])[:10] + " по " + str(list(dates.keys())[-1])[:10])
    #  Создание excel книги, в которую будут записываться данные
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    #  Заполнение заголовок столбцов
    sheet[f'A4'].value = "Дата"
    sheet[f'B4'].value = "Приход, шт"
    sheet[f'C4'].value = "Расход, шт"
    sheet[f'D4'].value = "Остаток паллет"
    sheet[f'A4'].fill = PatternFill(fill_type='solid', start_color='fffeee')
    sheet[f'B4'].fill = PatternFill(fill_type='solid', start_color='fffeee')
    sheet[f'C4'].fill = PatternFill(fill_type='solid', start_color='fffeee')
    sheet[f'D4'].fill = PatternFill(fill_type='solid', start_color='fffeee')

    excel_style(sheet, 4, thin_border)

    #  Создание объедененный ячеек для названия отчета
    sheet.merge_cells('A1:D1')
    sheet[f'A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A2:D2')
    sheet[f'A2'].alignment = Alignment(horizontal='center')

    sheet['A1'] = str(excel.iat[index_of_werehous_i, index_of_werehous_j])
    sheet['A2'] = "C " + str(list(dates.keys())[0])[:10] + " по " + str(list(dates.keys())[-1])[:10]

    #  Объединение ячеек для инфы по последнему паллету
    sheet.merge_cells('E3:G3')
    sheet[f'E3'].alignment = Alignment(horizontal='center')
    sheet['E3'] = "Последний паллет, шт"
    sheet[f'E3'].border = thin_border
    sheet[f'F3'].border = thin_border
    sheet[f'G3'].border = thin_border

    sheet.column_dimensions['E'].width = 13
    sheet.column_dimensions['F'].width = 13
    sheet.column_dimensions['G'].width = 13
    sheet[f'E4'].border = thin_border
    sheet[f'F4'].border = thin_border
    sheet[f'G4'].border = thin_border
    sheet[f'E4'].alignment = Alignment(horizontal='center')
    sheet[f'F4'].alignment = Alignment(horizontal='center')
    sheet[f'G4'].alignment = Alignment(horizontal='center')

    sheet[f'E4'].value = "650; 550; 500"
    sheet[f'F4'].value = "330/1; 330"
    sheet[f'G4'].value = "300; 200; 150"

    dates["end"] = index_of_end
    # print(list(dates.keys()))
    # print(list(dates.values()))
    # print()
    #  Далее будем считать количество паллетов;
    ind = 0
    ex_ind = 5

    while ind < len(list(dates.values())) - 1:
        # Начало и конец промежутка - индексы между датами
        start = list(dates.values())[ind]
        end = list(dates.values())[ind + 1]
        # print(f"Начало: {list(dates.keys())[ind]}")
        # print(f"Конец: {list(dates.keys())[ind + 1]}")
        sum_1 = 0
        sum_2 = 0
        sum_3 = 0
        is_product = False
        tmp = []
        expenses = 0
        profit = 0
        # Цикл между двумя датами;
        for i in range(start + 1, end):
            tmp.append(excel.iat[i, index_of_product_j])

            # print(excel.iat[i, index_of_product_j][:12])
            # print(excel.iat[i, 16])
            if not pd.Series(excel.iat[i, -3]).hasnans:
                profit += excel.iat[i, -3]
            # print(excel.iat[i, -3])
            if not pd.Series(excel.iat[i, -2]).hasnans:
                expenses += excel.iat[i, -2]
            if type(excel.iat[i, index_of_product_j]) == str and excel.iat[i, index_of_product_j][:12] == "Обогреватель":
                #  650, 550, 500
                if excel.iat[i, index_of_product_j][21:21 + 3] in articles_1 or excel.iat[i, index_of_product_j][21:21 + 5] in articles_1:
                    if not pd.Series(excel.iat[i, -1]).hasnans:
                        sum_1 += excel.iat[i, -1]
                #  330/1, 360
                if excel.iat[i, index_of_product_j][21:21 + 3] in articles_2 or excel.iat[i, index_of_product_j][21:21 + 5] in articles_2:
                    if not pd.Series(excel.iat[i, -1]).hasnans:
                        sum_2 += excel.iat[i, -1]
                #  150, 200, 300
                if excel.iat[i, index_of_product_j][21:21 + 3] in articles_3 or excel.iat[i, index_of_product_j][21:21 + 5] in articles_3:
                    if not pd.Series(excel.iat[i, -1]).hasnans:
                        sum_3 += excel.iat[i, -1]

        # print(f"Расходы: {expenses}")
        # print(f"Приход: {profit}")

        # print(pd.Series(tmp))
        for name in pd.Series(tmp):
            if name[:5] != "Обогр":
                # print(name[:5])
                # print("Обогреватель не найдет:", name)
                is_product = True
                break

        # if pd.Series(tmp).hasnans:
        #     is_product = True

        #  Обычные суммы товаров
        all_sums = [sum_1, sum_2, sum_3]
        # print(all_sums)
        #  Округленное количество паллетов
        if num[0] != 0:
            sum_1_ceil = math.ceil(sum_1 / num[0])
        else:
            sum_1_ceil = 0
        if num[1] != 0:
            sum_2_ceil = math.ceil(sum_2 / num[1])
        else:
            sum_2_ceil = 0
        if num[2] != 0:
            sum_3_ceil = math.ceil(sum_3 / num[2])
        else:
            sum_3_ceil = 0

        all_sums_ceil = [sum_1_ceil, sum_2_ceil, sum_3_ceil]
        # print(all_sums_ceil)

        last_palette = []
        #  Проверка на двойку
        for i in range(len(all_sums_ceil)):
            last_palette.append(all_sums[i] - (all_sums_ceil[i] - 1) * num[i])
            # print(last_palette)
            if last_palette[i] <= 2:
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].value = last_palette[i]
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].border = thin_border
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].alignment = Alignment(horizontal='center')
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='ee4723')
                # print(f"WARNING! В коллекции {i + 1} последний паллет = {last_palette[i]}")
                all_sums_ceil[i] -= 1
            else:
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].value = last_palette[i]
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].border = thin_border
                sheet[f'{chr(68 + i + 1)}{ex_ind}'].alignment = Alignment(horizontal='center')
                # print(f"В коллекции {i + 1} последний паллет = {last_palette[i]}")

        #  Получили количество паллетов
        S = sum(all_sums_ceil)
        if is_product:
            S += 1
            # print("!!!!! + 1 паллет !!!!!")

        # print("Кол-во Паллетов:", S)

        #  Формируем данные для записи в нужном виде:
        date1 = list(dates.keys())[ind][:10]
        date2 = list(dates.keys())[ind + 1][:10]
        if date2 == "end":
            date2 = date1
        d1 = date(int(date1[-4:]), int(date1[4]), int(date1[:2]))  # начальная дата
        d2 = date(int(date2[-4:]), int(date2[4]), int(date2[:2]))  # начальная дата
        delta_day = 1

        #  Запись данных в отчет
        sheet[f'A{ex_ind}'].value = str(d1)
        sheet[f'B{ex_ind}'].value = profit
        sheet[f'C{ex_ind}'].value = expenses
        sheet[f'D{ex_ind}'].value = S

        sheet[f'A{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='6abf40')
        sheet[f'B{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='66bfc5')
        sheet[f'C{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='66bfc5')
        sheet[f'D{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='f4af5a')

        excel_style(sheet, ex_ind, thin_border)

        #  Цикл для записи дат, которые входят в текущий промежуток
        while int(sheet[f'A{ex_ind}'].value[-2:]) != int(str(d2)[-2:]):
            ex_ind += 1
            sheet[f'A{ex_ind}'].value = str(d1 + timedelta(delta_day))
            sheet[f'B{ex_ind}'].value = ''
            sheet[f'C{ex_ind}'].value = ''
            sheet[f'E{ex_ind}'].value = ''
            sheet[f'F{ex_ind}'].value = ''
            sheet[f'G{ex_ind}'].value = ''
            sheet[f'D{ex_ind}'].value = S

            sheet[f'A{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='7cbf40')
            sheet[f'B{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='7cbfc5')
            sheet[f'C{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='7cbfc5')
            sheet[f'D{ex_ind}'].fill = PatternFill(fill_type='solid', start_color='e8af5a')

            excel_style(sheet, ex_ind, thin_border, 1)
            delta_day += 1
        ind += 1
    # print("Расчет окончен")
    wb.save(f'{result_name}')
    return msg1, msg2


#  Функция для отрисовки границ и выравнивания по центру
def excel_style(sheet, ex_ind, thin_border, check=0):
    sheet[f'A{ex_ind}'].border = thin_border
    sheet[f'B{ex_ind}'].border = thin_border
    sheet[f'C{ex_ind}'].border = thin_border
    sheet[f'D{ex_ind}'].border = thin_border

    sheet[f'A{ex_ind}'].alignment = Alignment(horizontal='center')
    sheet[f'B{ex_ind}'].alignment = Alignment(horizontal='center')
    sheet[f'C{ex_ind}'].alignment = Alignment(horizontal='center')
    sheet[f'D{ex_ind}'].alignment = Alignment(horizontal='center')
    if check == 1:
        sheet[f'E{ex_ind}'].border = thin_border
        sheet[f'F{ex_ind}'].border = thin_border
        sheet[f'G{ex_ind}'].border = thin_border
        sheet[f'E{ex_ind}'].alignment = Alignment(horizontal='center')
        sheet[f'F{ex_ind}'].alignment = Alignment(horizontal='center')
        sheet[f'G{ex_ind}'].alignment = Alignment(horizontal='center')


# Главная функция, из нее производится запуск
def start(path, result_name='Report'):
    # Считываем нужный файл
    excel = pd.read_excel(path)
    t_0 = time.time()
    msg1, msg2 = analyze_and_insert(excel, result_name)
    msg3 = "Расчет окончен"
    msg4 = f"Программа работала {round(time.time() - t_0, 2)} сек"
    #print(excel)
    return msg1, msg2, msg3, msg4



if __name__ == "__main__":
    start("Ланкс Новосиб.xls")







